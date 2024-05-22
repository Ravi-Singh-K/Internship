# from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active
# w1 = wb.create_sheet("My Sheet") # default i.e. at the end
# w2 = wb.create_sheet("My Sheet 2", 0)  # added at the initial position
# w3 = wb.create_sheet("My Sheet 3", -1)  # at the penultimate position
# ws.title = "New Title"
# w3 = wb['New Title']
# print(wb.sheetnames)
# for sheet in wb:
#     print(sheet.title)
# ws['A4'] = 4
# c = ws['A4']
# d = ws.cell(row=4, column=2, value=10)
# cell_range = ws['A1':'C2']
# print(cell_range)

# # for rows in ws.values():
# #     for value in rows:
# #         print(value)


# from openpyxl import Workbook
# from openpyxl.styles import Font

# wb = Workbook()
# ws = wb.active
# data = [['A','Apple','Axe'],['B','Ball','Bat'],['C','Canon','Cat'],['D','Doll','Dog']]

# for row in data:
#     ws.append(row)
# ft = Font(bold=True)
# for row in wb["A1:C1"]:
#     for cell in row:
#         cell.font = ft
# wb.save("Alphabet.xlsx")


# from openpyxl import Workbook,load_workbook

# wb = Workbook()
# ws = wb.active

# ws["A1"] = "Fruits"
# ws["A2"] = "Name"

# wb.save(filename="First_Sheet.xlsx")

# wb = load_workbook(filename="First_Sheet.xlsx")
# print(wb.sheetnames)
# sheet = wb.active
# print(sheet)
# print(sheet.title)
# print(sheet["A1"].value)
# print(sheet["A2"].value)

# from openpyxl import Workbook
# import json
# import requests

# url = requests.get('https://raw.githubusercontent.com/awasekhirni/jsondata/master/100books.json')
# data = json.loads(url.text)
# # print(data)
# key_list = []
# value_dict = {}
# for i in data:
#     for j, k in i.items():
#         # print(j,k)
#         if j not in key_list:
#             key_list.append(j)
#             for m in key_list:
#                 value_dict.setdefault(m)


# print(key_list)
# print(value_dict)

# print(dict_info)

# wb = Workbook()
# ws = wb.active


# wb.save(filename="100jsondata.xlsx")


from openpyxl import Workbook, load_workbook
import json
import requests

url = requests.get('https://raw.githubusercontent.com/awasekhirni/jsondata/master/100books.json')
json_data = json.loads(url.text)
# print(json_data)
dict_data = {}

for item in json_data:
    country = item.get('country')
    if country not in dict_data:
        dict_data[country] = []
    dict_data[country].append({k:v for k,v in item.items() if k!='country'})
# print(json.dumps(dict_data, indent=4))

wb=Workbook()
ws = wb.active

for country, data_value in dict_data.items():
    ws.append([country])
    for data in data_value:
        ws.append([v for v in data.values()])
wb.save("100booksortedjson.xlsx")



